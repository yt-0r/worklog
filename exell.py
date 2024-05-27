from openpyxl.styles.borders import Border, Side
import logging
from collections import defaultdict
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl import load_workbook
from process_JS import month_to_insert


#------------------------------------------------------------------------------
def insert_exel(config, df_total, wb, add:str=None, id_worker:int=None, year_now:str=None, month_now:dict=None, mode:str=None, df_base=None):
        # print('draw')
        if add == 'new':
            if month_now in wb.sheetnames: #удаляем лист который будем перезаписывать
                wb.remove(wb[month_now])
            ws = wb.create_sheet(month_now)  # создаём новый лист
            logging.info("Start insert exel NEW")
            # здесь смотреть нечего. проматываем примерно на 230-ю строку. тут тупо делается шапка и задаются стили на КАЖДУЮ ячейку! ЖЭСТЬ полная. бляпиздец...
            # сделал ошибку, что обращался к ячейкам по их имени. вдальнейшем буду стараться обращаться по индексу
            # ws.title = date_string  # так можно переименовать лист
            # замораживаем строки выше A7 что бы они не реагировали на прокрутку
            ws.freeze_panes = 'A7'
            # устанавливаем фильтры для полей 
            ws.auto_filter.ref = 'A6:G6'
            ws['B3'] = "ТАБЕЛЬ учета использования рабочего времени"
            ws['D3'] = month_now + ' ' + year_now
            ws['B3'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws['D3'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=18.0)

            ws.merge_cells('I1:AE1')
            ws.merge_cells('I2:AE2')
            ws.merge_cells('I3:AE3')
            ws.merge_cells('I4:AE4')
            thin_border = Border(left=Side(style='medium'),
                                  right=Side(style='medium'),
                                  top=Side(style='medium'),
                                  bottom=Side(style='medium'))
            ws.cell(row=4, column=9).border = thin_border
            for r in range(1, 5):
                ws.cell(row=r, column=8).border = Border(left=Side(style='medium'))
                ws.cell(row=r, column=31).border = Border(
                    right=Side(style='medium'))
            for c in range(8, 32):
                ws.cell(row=4, column=c).border = Border(
                    bottom=Side(style='medium'))
            ws.cell(row=4, column=8).border = Border(
                left=Side(style='medium'), bottom=Side(style='medium'))
            ws.cell(row=4, column=31).border = Border(
                right=Side(style='medium'), bottom=Side(style='medium'))

            ws['H1'] = "О"
            ws['H2'] = "Б"
            ws['H3'] = "В"
            ws['H4'] = "Н"
            ws['H1'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=16.0)
            ws['H1'].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['H2'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=16.0)
            ws['H2'].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['H3'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=16.0)
            ws['H3'].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['H4'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=16.0)
            ws['H4'].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            ws['I1'] = "Отпуск (заполняется только в белой строке)"
            ws['I2'] = "Больничный (заполняется только в белой строке)"
            ws['I3'] = "Выходной (заполняется только в белой строке)"
            ws['I4'] = "Неявка по невыясненным причинам  (заполняется только в белой строке)"
            ws['I1'].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['I2'].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['I3'].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws['I4'].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
    # ================ ставим морковку ============================================
            ws.merge_cells('AG1:AP4')
            
            for r in range(1, 5):
                ws.cell(row=r, column=33).border = Border(left=Side(style='medium'))
                ws.cell(row=r, column=42).border = Border(
                    right=Side(style='medium'))
            
            for c in range(33, 43):
                ws.cell(row=4, column=c).border = Border(
                    bottom=Side(style='medium'))
            
            ws.cell(row=4, column=33).border = Border(
                left=Side(style='medium'), bottom=Side(style='medium'))
            ws.cell(row=4, column=42).border = Border(
                right=Side(style='medium'), bottom=Side(style='medium'))
            
            
            
            ws['AG1'] = '"Морковным" цветом выделены дни, в которых была корректирока рабочего времени'
            ws['AG1'].fill = PatternFill(
                start_color="F79646", end_color="F79646", fill_type="solid")
            
            
            ws['AG1'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=False, sz=12.0)
            ws['AG1'].alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center')
            # ws.column_dimensions['C'].width = value = 32.57
            
    # =============================================================================

            ws['A6'] = "№"
            ws['B6'] = "Ф.И.О"
            ws['C6'] = "Подразделение"
            ws['D6'] = "специальность, профессия"
            ws['E6'] = "Контракт"
            ws['F6'] = "Н/Д"
            ws['G6'] = "Командировка"
            ws['A6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws['A6'].alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center')
            ws.column_dimensions['A'].width = value = 4.57

            ws['B6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws['B6'].alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center')
            ws.column_dimensions['B'].width = value = 32.57

            ws['C6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws['C6'].alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center')
            ws.column_dimensions['C'].width = value = 32.57

            ws['D6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws['D6'].alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center')
            ws.column_dimensions['D'].width = value = 22.14

            ws['E6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws['E6'].alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center')
            ws.column_dimensions['E'].width = value = 22.14

            ws['F6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['F'].width = value = 5
            ws['F6'].alignment = Alignment(
                textRotation=90, wrap_text=True, vertical='center', horizontal='center')

            ws['G6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['G'].width = value = 5
            ws['G6'].alignment = Alignment(
                textRotation=90, wrap_text=True, vertical='center', horizontal='center')

            ws.column_dimensions['H'].width = value = 3.7
            ws['H6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['I'].width = value = 3.7
            ws['I6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['J'].width = value = 3.7
            ws['J6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['K'].width = value = 3.7
            ws['K6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['L'].width = value = 3.7
            ws['L6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['M'].width = value = 3.7
            ws['M6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['N'].width = value = 3.7
            ws['N6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['O'].width = value = 3.7
            ws['O6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['P'].width = value = 3.7
            ws['P6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['Q'].width = value = 3.7
            ws['Q6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['R'].width = value = 3.7
            ws['R6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['S'].width = value = 3.7
            ws['S6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['T'].width = value = 3.7
            ws['T6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['U'].width = value = 3.7
            ws['U6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['V'].width = value = 3.7
            ws['V6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['W'].width = value = 3.7
            ws['W6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['X'].width = value = 3.7
            ws['X6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['Y'].width = value = 3.7
            ws['Y6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['Z'].width = value = 3.7
            ws['Z6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AA'].width = value = 3.7
            ws['AA6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AB'].width = value = 3.7
            ws['AB6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AC'].width = value = 3.7
            ws['AC6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AD'].width = value = 3.7
            ws['AD6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AE'].width = value = 3.7
            ws['AE6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AF'].width = value = 3.7
            ws['AF6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AG'].width = value = 3.7
            ws['AG6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AH'].width = value = 3.7
            ws['AH6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AI'].width = value = 3.7
            ws['AI6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AJ'].width = value = 3.7
            ws['AJ6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AK'].width = value = 3.7
            ws['AK6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AL'].width = value = 3.7
            ws['AL6'].font = openpyxl.styles.Font(
                name='Arial', charset=204, family=2.0, b=True, sz=12.0)

            ws['AM6'] = "Отработано смен (день)"
            ws['AN6'] = "Отработано смен (ночь)"
            ws['AO6'] = "Отпуск"
            ws['AP6'] = "Отгул"
            ws['AQ6'] = "Больничный"
            ws['AR6'] = "Неявка"
            ws['AS6'] = "Командировка"
            ws['AT6'] = "Отработано часы (день)"
            ws['AU6'] = "Отработано часы (ночь)"
            ws['AV6'] = "Отработано в выходные часы (день)"
            ws['AW6'] = "Отработано в выходные часы (ночь)"

            ws['AM6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AM'].width = value = 5
            ws['AM6'].alignment = Alignment(textRotation=90)
            ws['AM6'].fill = PatternFill(
                start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            ws['AN6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AN'].width = value = 5
            ws['AN6'].alignment = Alignment(textRotation=90)
            ws['AN6'].fill = PatternFill(
                start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            ws['AO6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AO'].width = value = 5
            ws['AO6'].alignment = Alignment(textRotation=90)
            ws['AO6'].fill = PatternFill(
                start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            ws['AP6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AP'].width = value = 5
            ws['AP6'].alignment = Alignment(textRotation=90)
            ws['AP6'].fill = PatternFill(
                start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            ws['AQ6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AQ'].width = value = 5
            ws['AQ6'].alignment = Alignment(textRotation=90)
            ws['AQ6'].fill = PatternFill(
                start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            ws['AR6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AR'].width = value = 5
            ws['AR6'].alignment = Alignment(textRotation=90)
            ws['AR6'].fill = PatternFill(
                start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

            ws['AS6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AS'].width = value = 5
            ws['AS6'].alignment = Alignment(textRotation=90)
            ws['AS6'].fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid")

            ws['AT6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AT'].width = value = 5
            ws['AT6'].alignment = Alignment(textRotation=90)
            ws['AT6'].fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid")

            ws['AU6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AU'].width = value = 5
            ws['AU6'].alignment = Alignment(textRotation=90)
            ws['AU6'].fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid")

            ws['AV6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AV'].width = value = 5
            ws['AV6'].alignment = Alignment(textRotation=90)
            ws['AV6'].fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid")
            
            ws['AW6'].font = openpyxl.styles.Font(
                name='Arial Cyr', charset=204, family=2.0, b=True, sz=12.0)
            ws.column_dimensions['AV'].width = value = 5
            ws['Aw6'].alignment = Alignment(textRotation=90)
            ws['AW6'].fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid")

            # ---------------- дрочево закончилось ------------------------------------
    
      # ("year_now " +str(arg['year'])+ " month_now " +str(arg['month'])
    
            # ---------- расставляем даты в экселе и красим выходные ----------
            days = df_base[(df_base['period_month'] == month_now) & 
                           (df_base['period_year'] == int(year_now))][['work_calendar_day', 'work_calendar_daytype']].drop_duplicates().to_dict('records')
            # print('daysrecords--->', days)
            for day in days:
                ws.cell(row=6, column=7+day['work_calendar_day'], value=day['work_calendar_day'])
                if day['work_calendar_daytype'] == 1:
                    ws.cell(row=6, column=7+day['work_calendar_day']).fill = PatternFill(
                            start_color="FF0000", end_color="FF0000", fill_type="solid")
            days = None
            day = None
            # -----------------------------------------------------------------
            
            logging.info("Stop insrt exel NEW")
        if add == 'add':
            logging.info("Start insert exel ADD "+str(id_worker))
            # ---------------- пишем все в эксель ------------------------------------
            # определяем последнюю строку с данными в экселе
            ws = wb[month_now]
                # date_event = date_string.split()
            last_empty_row = len(list(ws.rows))

            lr_last = last_empty_row
            
            start_lr_last = last_empty_row # в start_lr_last сохраняем номер строки с которой пишем новоо сотрудника. необходимо для расчета формул
            # --------- заполняем Excel данными из DataFrame ------------------
# ================================================================================================
# ВНИМАНИЕ!!!!!!! тут написана хуйня!!!!!!
# надо что бы табель на человека заполнялся ПОСТРОЧНО!!! иначе будет лажа!!!!
# к примеру, сколько дней в месяце, столько и будет общехозяйственных кнтрактов
# вначале берем человека, в нем контракт и проставляем ему все дни и ни как иначе!!!!
# вроде, поправил хуйню, но это не точно )))
# ================================================================================================
            # --- все переписываем -----
            # -- все равно пишу хуйню. дублирую код. лишь бы сдать. потом надо все привести в чувства----
            # print('id_worker-->>', id_worker)
            # print('month_now-->>', month_now)
            # print('year_now-->>', year_now)
            # print('type year_now-->>', type(year_now))
            # print('df_base--->>', df_base)
            
            days = df_base[(df_base['period_month'] == month_now) & 
                           (df_base['period_year'] == int(year_now))
                           ][['work_calendar_day', 'work_calendar_daytype', 'period_month']].drop_duplicates().to_dict('records')
            # print('days-->>', days)
            
            
            # # =====================================================================
            # # ---------- пишем формулу для вычисления выходных дней ----------------
            # formula_weekend_day = '=SUM('
            # for day in days:
            #     formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last+k, column=s+2).coordinate + ','
            
            # formula_weekend_day = formula_weekend_day.rstrip('+')
            # formula_weekend_day = formula_weekend_day + ')'
            # # ======================================================================
            
            
            job_description = df_base[['job_name', 'job_department', 'job_position']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now))].drop_duplicates().to_dict()
            # print('job_description first-->>', job_description)
            # print('id_worker job_description---->',id_worker)
            # print('month_now job_description---->',month_now)
            # print('year_now job_description---->',int(year_now))
            # ниже в поиск добавить к имеи контракта день/ночь
            # -- выбираем все уникальные контракты в этом году и этом месяце на этого работника ---
            kontrakts = df_base['kontrakt_name'].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now))].drop_duplicates().to_dict()
            # print('search-->>', df_base['kontrakt_name'].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & ((df_base['kontrakt_name'] != 'Общехозяйственный') & (df_base['kontrakt_name'] != 'Офис'))])
            # print('kontrakts-->>', kontrakts)
            # тут мы должны напсать что-то вроде того, что типа первой строкой в каждой фамилии должен быть родной контракт и именно shift/correction
            # потом мы делим строки на несколько блоков:
                # -все шифты и корректировки по всем котрактам
                # -все переработки по всем контрактам в день
                # -все переработки по всем контрактам в ночь
                # -все командировки по всем контрактам (эти три пункта обрабатываем в одном цикле. важно, что бы эти эвенты были сортированы по контрактам и находились один под другим)
            # что делать со всякими больничными пока не ясно надо обсудить, скорее всего надо будет на первом шаге 
            # искать все эти эвенты (т.к. они будут только на основной контракт ставиться) и дополнительно их парсить
            
            
            # здесь втакой момент. нам нужно как-то запомнить номер строки с которой начали заполнять очередного работника
            # нужно по ходу обработки заполнять словарь вида {"номер дня":["координата 1","координата 2", "координата n"]}
            # + плюс такой же для ночи
            formula_working_shift_day = defaultdict(list)
            formula_working_shift_night = defaultdict(list)
            # ищем и обрабатываем основной контракт
            base_kontrakt_1 = df_base['kontrakt_name'].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Общехозяйственный') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates().to_dict()
            # print('base_kontrakt_1-->>', base_kontrakt_1)
            # print('len base_kontrakt_1-->>', len(base_kontrakt_1))
            # start_lr_last = 0
            if len(base_kontrakt_1) > 0:
                # здесь Общехозяйственный в день
                if len(df_base[['day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Общехозяйственный') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital')) & (df_base['day_night'] == 'day')].drop_duplicates()) > 0:
                    # start_lr_last = lr_last # запоминаем эту строку как первую
                    lr_last = lr_last +1
                    
                    for day in days:
                        ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                        ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                        ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                        ws.cell(row=lr_last, column=5, value=list(base_kontrakt_1.values())[0])
                        ws.cell(row=lr_last, column=6, value='День')
                        
                        stata_in_day = df_base[['kontrakt_timetracking', 'kontrakt_filter', 'event', 'work_calendar_daytype']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Общехозяйственный') & (df_base['day_night'] == 'day') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates().to_dict()
                        # print('stata_in_day-->>', stata_in_day)
                        if len(df_base[['kontrakt_timetracking', 'kontrakt_filter', 'event']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Общехозяйственный') & (df_base['day_night'] == 'day') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates()) > 0:
                            # print('stata_in_day 1-->>', list(stata_in_day['kontrakt_timetracking'].values())[0])
                            
                            if list(stata_in_day['event'].values())[0] == 'otpusk':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='О')
                            
                            if list(stata_in_day['event'].values())[0] == 'compensatory':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='От')
                                # print('12121212')
                            
                            if list(stata_in_day['event'].values())[0] == 'hospital':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='Б')

                            if list(stata_in_day['event'].values())[0] == 'shift':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='')
                                if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                                    if list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                    
                            if list(stata_in_day['event'].values())[0] == 'correction':
                                
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='Н',)
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day']).fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
                                    
                                    
                                if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                                    if list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                            
                            if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
                            
                            formula_working_shift_day[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                            
                            # formula_working_shift_day[day['work_calendar_day']].add(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                    # print(stata_in_day)
                    # print('formula_working_shift_day-->>', formula_working_shift_day)
                    # отработано смен (день)
                    ws.cell(row=lr_last, column=39).value = '=COUNT('+ws.cell(row=lr_last,
                                                                                column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                    ws.cell(row=lr_last, column=39).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # отработано часы (день)
                    ws.cell(row=lr_last, column=46).value = '=SUM('+ws.cell(row=lr_last,
                                                                              column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                    ws.cell(row=lr_last, column=46).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    
                    
                    
                    # отпуск
                    ws.cell(row=lr_last, column=41).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"О")'''
                    ws.cell(row=lr_last, column=41).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # отгул
                    ws.cell(row=lr_last, column=42).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"От")'''
                    ws.cell(row=lr_last, column=42).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # больничный
                    ws.cell(row=lr_last, column=43).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"Б")'''
                    ws.cell(row=lr_last, column=43).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # неявка
                    ws.cell(row=lr_last, column=44).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"Н")'''
                    ws.cell(row=lr_last, column=44).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    
                    # отработано в выходные
                    # =====================================================================
                    # ---------- пишем формулу для вычисления выходных дней в день----------------
                    formula_weekend_day = '=SUM('
                    
                    # print('days-->>', days)
                    for day in days:
                        if day['work_calendar_daytype'] == 1:
                            formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day['work_calendar_day']+7).coordinate + ','
                    # print('formula_weekend_day-->', formula_weekend_day)
                    formula_weekend_day = formula_weekend_day.rstrip(',')
                    formula_weekend_day = formula_weekend_day + ')'
                    # print('formula_weekend_day-->', formula_weekend_day)
                    if formula_weekend_day != '=SUM()':
                        ws.cell(row=lr_last, column=48).value = formula_weekend_day
                        ws.cell(row=lr_last, column=48).font = openpyxl.styles.Font(
                                name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # ======================================================================
                    
                # здесь Общехозяйственный в ночь
                if len(df_base[['day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Общехозяйственный') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital')) & (df_base['day_night'] == 'night')].drop_duplicates()) > 0:
                    lr_last = lr_last +1
                    
                    for day in days:
                        # print('day-->>', day)
                        # print('day 1-->>', day['work_calendar_day'])
                        # print('job-->>', list(job_description['job_name'].values())[0])
                        ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                        ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                        ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                        ws.cell(row=lr_last, column=5, value=list(base_kontrakt_1.values())[0])
                        ws.cell(row=lr_last, column=6, value='Ночь')
                        
                        stata_in_day = df_base[['kontrakt_timetracking', 'kontrakt_filter', 'event', 'work_calendar_daytype']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Общехозяйственный') & (df_base['day_night'] == 'night') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates().to_dict()
                        # print('stata_in_day-->>', stata_in_day)
                        if len(df_base[['kontrakt_timetracking', 'kontrakt_filter', 'event']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Общехозяйственный') & (df_base['day_night'] == 'night') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates()) > 0:
                            # print('stata_in_day 1-->>', list(stata_in_day['kontrakt_timetracking'].values())[0])
                            
                            if list(stata_in_day['event'].values())[0] == 'otpusk':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='О')
                            
                            if list(stata_in_day['event'].values())[0] == 'compensatory':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='От')
                            
                            
                            if list(stata_in_day['event'].values())[0] == 'hospital':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='Б')
                            if list(stata_in_day['event'].values())[0] == 'shift':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='')
                                if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                                    if list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                    
                            if list(stata_in_day['event'].values())[0] == 'correction':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='Н')
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day']).fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
                                if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                                    if list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                            
                            if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
                                
                            formula_working_shift_night[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                
                    # отработано смен (ночь)
                    ws.cell(row=lr_last, column=40).value = '=COUNT('+ws.cell(row=lr_last,
                                                                                column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                    ws.cell(row=lr_last, column=40).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # отработано часов (ночь)
                    ws.cell(row=lr_last, column=47).value = '=SUM('+ws.cell(row=lr_last,
                            column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                    ws.cell(row=lr_last, column=47).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                
                    # отпуск
                    ws.cell(row=lr_last, column=41).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"О")'''
                    ws.cell(row=lr_last, column=41).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # отгул
                    ws.cell(row=lr_last, column=42).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"От")'''
                    ws.cell(row=lr_last, column=42).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # больничный
                    ws.cell(row=lr_last, column=43).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"Б")'''
                    ws.cell(row=lr_last, column=43).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # неявка
                    ws.cell(row=lr_last, column=44).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"Н")'''
                    ws.cell(row=lr_last, column=44).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    
                    # отработано в выходные
                    # =====================================================================
                    # ---------- пишем формулу для вычисления выходных дней в ночь----------------
                    formula_weekend_day = '=SUM('
                    
                    # print('days-->>', days)
                    for day in days:
                        if day['work_calendar_daytype'] == 1:
                            formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day['work_calendar_day']+7).coordinate + ','
                    # print('formula_weekend_day-->', formula_weekend_day)
                    formula_weekend_day = formula_weekend_day.rstrip(',')
                    formula_weekend_day = formula_weekend_day + ')'
                    # print('formula_weekend_day-->', formula_weekend_day)
                    if formula_weekend_day != '=SUM()':
                        ws.cell(row=lr_last, column=49).value = formula_weekend_day
                        ws.cell(row=lr_last, column=49).font = openpyxl.styles.Font(
                                name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # ======================================================================
                
                
            base_kontrakt_2 = df_base['kontrakt_name'].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Офис') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates().to_dict()

            # print('base_kontrakt_2-->>', base_kontrakt_2)
            # print('len base_kontrakt_2-->>', len(base_kontrakt_2))
            if len(base_kontrakt_2) > 0:
                # здесь Офис в день
                # if start_lr_last == 0:
                #     start_lr_last = lr_last # запоминаем эту строку как первую
                if len(df_base[['day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Офис') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital')) & (df_base['day_night'] == 'day')].drop_duplicates()) > 0:
                    lr_last = lr_last +1
                    for day in days:
                        ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                        ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                        ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                        ws.cell(row=lr_last, column=5, value=list(base_kontrakt_2.values())[0])
                        ws.cell(row=lr_last, column=6, value='День')
                        
                        stata_in_day = df_base[['kontrakt_timetracking', 'kontrakt_filter', 'event', 'work_calendar_daytype']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Офис') & (df_base['day_night'] == 'day') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates().to_dict()
                        # print('stata_in_day-->>', stata_in_day)
                        if len(df_base[['kontrakt_timetracking', 'kontrakt_filter', 'event']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Офис') & (df_base['day_night'] == 'day') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates()) > 0:
                            # print('stata_in_day 1-->>', list(stata_in_day['kontrakt_timetracking'].values())[0])
                            
                            if list(stata_in_day['event'].values())[0] == 'otpusk':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='О')
                            
                            if list(stata_in_day['event'].values())[0] == 'compensatory':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='От')
                            
                            if list(stata_in_day['event'].values())[0] == 'hospital':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='Б')
                            if list(stata_in_day['event'].values())[0] == 'shift':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='')
                                if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                                    if list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                    
                            if list(stata_in_day['event'].values())[0] == 'correction':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='Н')
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day']).fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
                                if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                                    if list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                            
                            if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
                            
                            
                            # if id_worker == 103:
                            #     print('kontrakt-->>', base_kontrakt_2)
                            #     print('stats--->>', stata_in_day)
                            #     # print('base-->>>', df_base['kontrakt_name'].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Офис') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].to_dict())
                            
                            formula_working_shift_day[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))


                    # отработано смен (день)
                    ws.cell(row=lr_last, column=39).value = '=COUNT('+ws.cell(row=lr_last,
                                                                                column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                    ws.cell(row=lr_last, column=39).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # отработано часы (день)
                    ws.cell(row=lr_last, column=46).value = '=SUM('+ws.cell(row=lr_last,
                                                                              column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                    ws.cell(row=lr_last, column=46).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    
                    
                    
                    # отпуск
                    ws.cell(row=lr_last, column=41).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"О")'''
                    ws.cell(row=lr_last, column=41).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # отгул
                    ws.cell(row=lr_last, column=42).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"От")'''
                    ws.cell(row=lr_last, column=42).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # больничный
                    ws.cell(row=lr_last, column=43).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"Б")'''
                    ws.cell(row=lr_last, column=43).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # неявка
                    ws.cell(row=lr_last, column=44).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"Н")'''
                    ws.cell(row=lr_last, column=44).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    
                    # отработано в выходные
                    # =====================================================================
                    # ---------- пишем формулу для вычисления выходных дней в день----------------
                    formula_weekend_day = '=SUM('
                    
                    # print('days-->>', days)
                    for day in days:
                        if day['work_calendar_daytype'] == 1:
                            formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day['work_calendar_day']+7).coordinate + ','
                    # print('formula_weekend_day-->', formula_weekend_day)
                    formula_weekend_day = formula_weekend_day.rstrip(',')
                    formula_weekend_day = formula_weekend_day + ')'
                    # print('formula_weekend_day-->', formula_weekend_day)
                    if formula_weekend_day != '=SUM()':
                        ws.cell(row=lr_last, column=48).value = formula_weekend_day
                        ws.cell(row=lr_last, column=48).font = openpyxl.styles.Font(
                                name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # ======================================================================
                
                # здесь Офис в ночь

                if len(df_base[['day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Офис') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital')) & (df_base['day_night'] == 'night')].drop_duplicates()) > 0:
                    lr_last = lr_last +1
                    for day in days:
                        ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                        ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                        ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                        ws.cell(row=lr_last, column=5, value=list(base_kontrakt_2.values())[0])
                        ws.cell(row=lr_last, column=6, value='Ночь')
                        
                        stata_in_day = df_base[['kontrakt_timetracking', 'kontrakt_filter', 'event', 'work_calendar_daytype']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Офис') & (df_base['day_night'] == 'night') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates().to_dict()
                        # print('stata_in_day-->>', stata_in_day)
                        if len(df_base[['kontrakt_timetracking', 'kontrakt_filter', 'event']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == 'Офис') & (df_base['day_night'] == 'night') & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction') | (df_base['event'] == 'otpusk') | (df_base['event'] == 'compensatory') | (df_base['event'] == 'hospital'))].drop_duplicates()) > 0:
                            # print('stata_in_day 1-->>', list(stata_in_day['kontrakt_timetracking'].values())[0])
                            
                            if list(stata_in_day['event'].values())[0] == 'otpusk':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='О')
                            
                            if list(stata_in_day['event'].values())[0] == 'compensatory':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='От')
                            
                            
                            if list(stata_in_day['event'].values())[0] == 'otpusk':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='От')
                            
                            if list(stata_in_day['event'].values())[0] == 'hospital':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='Б')
                            if list(stata_in_day['event'].values())[0] == 'shift':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='')
                                if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                                    if list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                    
                            if list(stata_in_day['event'].values())[0] == 'correction':
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                                if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='Н')
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day']).fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
                                if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                                    if list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                            
                            if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
                
                            formula_working_shift_night[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))


                    # отработано смен (ночь)
                    ws.cell(row=lr_last, column=40).value = '=COUNT('+ws.cell(row=lr_last,
                                                                                column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                    ws.cell(row=lr_last, column=40).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # отработано часов (ночь)
                    ws.cell(row=lr_last, column=47).value = '=SUM('+ws.cell(row=lr_last,
                            column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                    ws.cell(row=lr_last, column=47).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    

                    # отпуск
                    ws.cell(row=lr_last, column=41).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"О")'''
                    ws.cell(row=lr_last, column=41).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # отгул
                    ws.cell(row=lr_last, column=42).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"От")'''
                    ws.cell(row=lr_last, column=42).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # больничный
                    ws.cell(row=lr_last, column=43).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"Б")'''
                    ws.cell(row=lr_last, column=43).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # неявка
                    ws.cell(row=lr_last, column=44).value = '=COUNTIF('+ws.cell(row=lr_last,
                                                                                  column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+''',"Н")'''
                    ws.cell(row=lr_last, column=44).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    
                    # отработано в выходные
                    # =====================================================================
                    # ---------- пишем формулу для вычисления выходных дней в ночь----------------
                    formula_weekend_day = '=SUM('
                    
                    # print('days-->>', days)
                    for day in days:
                        if day['work_calendar_daytype'] == 1:
                            formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day['work_calendar_day']+7).coordinate + ','
                    # print('formula_weekend_day-->', formula_weekend_day)
                    formula_weekend_day = formula_weekend_day.rstrip(',')
                    formula_weekend_day = formula_weekend_day + ')'
                    # print('formula_weekend_day-->', formula_weekend_day)
                    if formula_weekend_day != '=SUM()':
                        ws.cell(row=lr_last, column=49).value = formula_weekend_day
                        ws.cell(row=lr_last, column=49).font = openpyxl.styles.Font(
                                name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # ======================================================================
                
            # -- теперь выводим все шифты и корректировки по всем котрактам
            for kontrakt in list(kontrakts.values()):
                # print('df_base-->>', df_base)
                # что бы не выводились пустые строки в excel, нужно проверять есть ли хоть что-то на этого сотрудника в этом месяце и годе на этом контракте
                if len(df_base.loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction')) & (((df_base['kontrakt_name'] != 'Общехозяйственный') & (df_base['kontrakt_name'] != 'Офис')))]) >0:
                    lr_last = lr_last +1
                    for day in days:
                        stata_in_day = df_base[['kontrakt_timetracking', 'kontrakt_filter', 'day_night', 'work_calendar_daytype', 'work_time']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & ((df_base['event'] == 'shift') | (df_base['event'] == 'correction')) & (((df_base['kontrakt_name'] != 'Общехозяйственный') & (df_base['kontrakt_name'] != 'Офис')))].drop_duplicates().to_dict()
                        # print('day-->>', day)
                        # print('stata-->>', stata_in_day['kontrakt_timetracking'])
                        if len(stata_in_day['kontrakt_timetracking']) > 0:
                            # lr_last = lr_last +1
                            ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                            ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                            ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                            ws.cell(row=lr_last, column=5, value=kontrakt)
                            
                            if list(stata_in_day['day_night'].values())[0] == 'day':
                                ws.cell(row=lr_last, column=6, value='День')
                                
                                
                                # отработано смен (день)
                                ws.cell(row=lr_last, column=39).value = '=COUNT('+ws.cell(row=lr_last,
                                                                                            column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=39).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # отработано часы (день)
                                ws.cell(row=lr_last, column=46).value = '=SUM('+ws.cell(row=lr_last,
                                                                                          column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=46).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                
                                # отработано в выходные
                                # =====================================================================
                                # ---------- пишем формулу для вычисления выходных дней в день----------------
                                formula_weekend_day = '=SUM('
                                
                                # print('days-->>', days)
                                for day1 in days:
                                    if day1['work_calendar_daytype'] == 1:
                                        formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day1['work_calendar_day']+7).coordinate + ','
                                # print('formula_weekend_day-->', formula_weekend_day)
                                formula_weekend_day = formula_weekend_day.rstrip(',')
                                formula_weekend_day = formula_weekend_day + ')'
                                # print('formula_weekend_day-->', formula_weekend_day)
                                if formula_weekend_day != '=SUM()':
                                    ws.cell(row=lr_last, column=48).value = formula_weekend_day
                                    ws.cell(row=lr_last, column=48).font = openpyxl.styles.Font(
                                            name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # ======================================================================
                                
                                formula_working_shift_day[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                                # formula_working_shift_day[day['work_calendar_day']].add(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                            if list(stata_in_day['day_night'].values())[0] == 'night':
                                ws.cell(row=lr_last, column=6, value='Ночь')
                                # print('stata_in_day_higth-->>', stata_in_day)
                                # print('id_worker-->>', id_worker)
                                # отработано смен (ночь)
                                ws.cell(row=lr_last, column=40).value = '=COUNT('+ws.cell(row=lr_last,
                                                                                            column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=40).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # отработано часов (ночь)
                                ws.cell(row=lr_last, column=47).value = '=SUM('+ws.cell(row=lr_last,
                                        column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=47).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                
                                # отработано в выходные
                                # =====================================================================
                                # ---------- пишем формулу для вычисления выходных дней в ночь----------------
                                formula_weekend_day = '=SUM('
                                
                                # print('days-->>', days)
                                for day2 in days:
                                    if day2['work_calendar_daytype'] == 1:
                                        formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day2['work_calendar_day']+7).coordinate + ','
                                # print('formula_weekend_day-->', formula_weekend_day)
                                formula_weekend_day = formula_weekend_day.rstrip(',')
                                formula_weekend_day = formula_weekend_day + ')'
                                # print('formula_weekend_day-->', formula_weekend_day)
                                if formula_weekend_day != '=SUM()':
                                    ws.cell(row=lr_last, column=49).value = formula_weekend_day
                                    ws.cell(row=lr_last, column=49).font = openpyxl.styles.Font(
                                            name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # ======================================================================
                                
                                formula_working_shift_night[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                        
                            ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                            if list(stata_in_day['kontrakt_timetracking'].values())[0] == 0:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='')
                            # print('-->>', list(stata_in_day['work_calendar_daytype'].values())[0])
                            if list(stata_in_day['work_calendar_daytype'].values())[0] == 0:
                                if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
                            if (list(stata_in_day['work_calendar_daytype'].values())[0] == 1) & (list(stata_in_day['work_time'].values())[0] > 0):
                                if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                    ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
                                    # if id_worker == 103:
                                        # print('kontrakt-->>', base_kontrakt_2)
                                        # print('stata--->>', stata_in_day)
                            
                            
            # -----------------------------------------------------------------
            # -- теперь выставляем переработки вместе с командировками -------
            # print('kontrakts-->>', kontrakts)
            for kontrakt in list(kontrakts.values()):
                if len(df_base[['day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & (df_base['event'] == 'permit') & (df_base['day_night'] == 'day')].drop_duplicates()) > 0:
                    lr_last = lr_last +1
                    for day in days:
                        # здесь переработки в день
                        stata_in_day = df_base[['kontrakt_timetracking', 'kontrakt_filter', 'day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & (df_base['event'] == 'permit') & (df_base['day_night'] == 'day')].drop_duplicates().to_dict()
                        # if id_worker == 114:
                            # print('day-->>', day)
                            # print('kontrakt-->>', kontrakt)
                            # print('stata-->>', stata_in_day)
                        
                        if len(stata_in_day['kontrakt_timetracking']) > 0:
                            
                            ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                            ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                            ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                            ws.cell(row=lr_last, column=5, value=kontrakt)
                            
                            if list(stata_in_day['day_night'].values())[0] == 'day':
                                # if id_worker == 114:
                                #     print('-->>', list(stata_in_day['day_night'].values())[0])
                                    # print('kontrakt-->>', kontrakt)
                                    # print('stata-->>', stata_in_day)
                                ws.cell(row=lr_last, column=6, value='День')
                                
                                
                                # отработано смен (день)
                                ws.cell(row=lr_last, column=39).value = '=COUNT('+ws.cell(row=lr_last,
                                                                                            column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=39).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # отработано часы (день)
                                ws.cell(row=lr_last, column=46).value = '=SUM('+ws.cell(row=lr_last,
                                                                                          column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=46).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                
                                # отработано в выходные
                                # =====================================================================
                                # ---------- пишем формулу для вычисления выходных дней в день----------------
                                formula_weekend_day = '=SUM('
                                
                                # print('days-->>', days)
                                for day_hol in days:
                                    if day_hol['work_calendar_daytype'] == 1:
                                        formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day_hol['work_calendar_day']+7).coordinate + ','
                                # print('formula_weekend_day-->', formula_weekend_day)
                                formula_weekend_day = formula_weekend_day.rstrip(',')
                                formula_weekend_day = formula_weekend_day + ')'
                                # print('formula_weekend_day-->', formula_weekend_day)
                                if formula_weekend_day != '=SUM()':
                                    ws.cell(row=lr_last, column=48).value = formula_weekend_day
                                    ws.cell(row=lr_last, column=48).font = openpyxl.styles.Font(
                                            name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # ======================================================================
                                
                                formula_working_shift_day[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                                # formula_working_shift_day[day['work_calendar_day']].add(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
# =============================================================================
# ХЗ зачем это написано. но оно написано и все как-будто бы работало, но странно это все ))....
#                             if list(stata_in_day['day_night'].values())[0] == 'night':
#                                 ws.cell(row=lr_last, column=6, value='Ночь')
# =============================================================================
                        
                            ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                            if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
                    # здесь переработки в ночь
                if len(df_base[['day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & (df_base['event'] == 'permit') & (df_base['day_night'] == 'night')].drop_duplicates()) > 0:
                    lr_last = lr_last +1
                    # print('search night-->>', len(df_base[['day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & (df_base['event'] == 'permit') & (df_base['day_night'] == 'night')].drop_duplicates()))#.to_dict())
                    for day in days:
                        stata_in_day = df_base[['kontrakt_timetracking', 'kontrakt_filter', 'day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & (df_base['event'] == 'permit') & (df_base['day_night'] == 'night')].drop_duplicates().to_dict()
                        if len(stata_in_day['kontrakt_timetracking']) > 0:
                            
                            ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                            ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                            ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                            ws.cell(row=lr_last, column=5, value=kontrakt)
                            
                            # if list(stata_in_day['day_night'].values())[0] == 'day':
                            #     ws.cell(row=lr_last, column=6, value='День')
                            if list(stata_in_day['day_night'].values())[0] == 'night':
                                ws.cell(row=lr_last, column=6, value='Ночь')
                                
                                
                                # отработано смен (ночь)
                                ws.cell(row=lr_last, column=40).value = '=COUNT('+ws.cell(row=lr_last,
                                                                                            column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=40).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # отработано часов (ночь)
                                ws.cell(row=lr_last, column=47).value = '=SUM('+ws.cell(row=lr_last,
                                        column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=47).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                
                                # отработано в выходные
                                # =====================================================================
                                # ---------- пишем формулу для вычисления выходных дней в ночь----------------
                                formula_weekend_day = '=SUM('
                                
                                # print('days-->>', days)
                                for day_hol in days:
                                    if day_hol['work_calendar_daytype'] == 1:
                                        formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day_hol['work_calendar_day']+7).coordinate + ','
                                # print('formula_weekend_day-->', formula_weekend_day)
                                formula_weekend_day = formula_weekend_day.rstrip(',')
                                formula_weekend_day = formula_weekend_day + ')'
                                # print('formula_weekend_day-->', formula_weekend_day)
                                if formula_weekend_day != '=SUM()':
                                    ws.cell(row=lr_last, column=49).value = formula_weekend_day
                                    ws.cell(row=lr_last, column=49).font = openpyxl.styles.Font(
                                            name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # ======================================================================
                                
                        
                            ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                            if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
                                
                            formula_working_shift_night[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                            
                    # здесь командировки
                if len(df_base[['day_night']].loc[(df_base['id_worker'] == id_worker) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & (df_base['event'] == 'trip')].drop_duplicates()) > 0:
                    lr_last = lr_last +1
                    for day in days:
                        stata_in_day = df_base[['kontrakt_timetracking', 'kontrakt_filter', 'day_night', 'event', 'work_calendar_daytype']].loc[(df_base['id_worker'] == id_worker) & (df_base['work_calendar_day'] == day['work_calendar_day']) & (df_base['period_month'] == month_now) & (df_base['period_year'] == int(year_now)) & (df_base['kontrakt_name'] == kontrakt) & (df_base['event'] == 'trip')].drop_duplicates().to_dict()
                        if len(stata_in_day['kontrakt_timetracking']) > 0:
                            
                            ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                            ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                            ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                            ws.cell(row=lr_last, column=5, value=kontrakt)
                            
                            if list(stata_in_day['day_night'].values())[0] == 'day':
                                ws.cell(row=lr_last, column=6, value='День')
                                
                                # формулу, которая считает смены в командировках делим на 8 потому что командировки оформляют на несколько контрактов. 
                                # в первоисточнике в этой строке стояла формула =COUNT(H263:AL263), теперь =SUM(H263:AL263)
                                # отработано смен (день)
                                ws.cell(row=lr_last, column=39).value = '=SUM('+ws.cell(row=lr_last,
                                                                                            column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')/8'
                                ws.cell(row=lr_last, column=39).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # отработано часов (день)
                                ws.cell(row=lr_last, column=46).value = '=SUM('+ws.cell(row=lr_last,
                                        column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=46).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                
                                formula_working_shift_day[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                                # formula_working_shift_day[day['work_calendar_day']].add(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                            if list(stata_in_day['day_night'].values())[0] == 'night':
                                ws.cell(row=lr_last, column=6, value='Ночь')
                                
                                # отработано смен (ночь)
                                ws.cell(row=lr_last, column=40).value = '=SUM('+ws.cell(row=lr_last,
                                                                                            column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')/8'
                                ws.cell(row=lr_last, column=40).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # отработано часов (ночь)
                                ws.cell(row=lr_last, column=47).value = '=SUM('+ws.cell(row=lr_last,
                                        column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')'
                                ws.cell(row=lr_last, column=47).font = openpyxl.styles.Font(
                                    name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                
                                # отработано в выходные
                                # =====================================================================
                                # ---------- пишем формулу для вычисления выходных дней в день----------------
                                formula_weekend_day = '=SUM('
                                
                                # print('days-->>', days)
                                for day in days:
                                    if day['work_calendar_daytype'] == 1:
                                        formula_weekend_day = formula_weekend_day + ws.cell(row=lr_last, column=day['work_calendar_day']+7).coordinate + ','
                                # print('formula_weekend_day-->', formula_weekend_day)
                                formula_weekend_day = formula_weekend_day.rstrip(',')
                                formula_weekend_day = formula_weekend_day + ')'
                                # print('formula_weekend_day-->', formula_weekend_day)
                                if formula_weekend_day != '=SUM()':
                                    ws.cell(row=lr_last, column=48).value = formula_weekend_day
                                    ws.cell(row=lr_last, column=48).font = openpyxl.styles.Font(
                                            name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                # ======================================================================
                                formula_working_shift_night[day['work_calendar_day']].append(str(ws.cell(row=lr_last, column=7+day['work_calendar_day']).coordinate))
                                
                            # командировка
                            # здесь то же каунт заменили на сумм и поделили на 8 из тех же соображений
                            ws.cell(row=lr_last, column=45).value = '=COUNTIF('+ws.cell(row=lr_last, column=8).coordinate+':'+ws.cell(row=lr_last,
                                    column=38).coordinate+''',"В")''' + '+SUM('+ws.cell(row=lr_last, column=8).coordinate+':'+ws.cell(row=lr_last, column=38).coordinate+')/8'
                            ws.cell(row=lr_last, column=45).font = openpyxl.styles.Font(
                                name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                                
                            if list(stata_in_day['event'].values())[0] == 'trip':
                                ws.cell(row=lr_last, column=7, value='К-ка')
                            # print('lr_last-->>', lr_last)
                            # print('stata_in_day-->>', stata_in_day['kontrakt_timetracking'])
                            # print('day-->>', day['work_calendar_day'])
                            ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(stata_in_day['kontrakt_timetracking'].values())[0])
                            if list(stata_in_day['work_calendar_daytype'].values())[0] == 1:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day'], value='В')
                            if list(stata_in_day['kontrakt_filter'].values())[0] != '0' and list(stata_in_day['kontrakt_timetracking'].values())[0] > 0:
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).hyperlink = config["jira"]["jira_server"]+'/issues/?jql=issue in (' + list(stata_in_day['kontrakt_filter'].values())[0] + ')'
                                ws.cell(row=lr_last, column=7+day['work_calendar_day']).style = "Hyperlink"
            # ---------------------------------------------------------
            # --- ставим строку Total -------------------------------------
            # print('===================')
            # print('lr_last-->>', lr_last)
            # print('job_description len-->>', len(job_description['job_name']))
            # print('day-->>', day['work_calendar_day'])
            # print('formula_working_shift_day-->', formula_working_shift_day)
            if len(job_description['job_name']) >0:
                lr_last = lr_last +1
                ws.cell(row=lr_last, column=2, value=list(job_description['job_name'].values())[0])
                ws.cell(row=lr_last, column=3, value=list(job_description['job_department'].values())[0])
                ws.cell(row=lr_last, column=4, value=list(job_description['job_position'].values())[0])
                ws.cell(row=lr_last, column=5, value='ИТОГО:')
            
            
            
                #  здесь ИТОГО "отработано смен в день/ночь". 
                ## по непонятной причине set(formula_working_shift_day[fday]) 
                ## портит словарь странным образом - часть данных остается словарем, а часть данных становится списком из одного элемента.
                ## по идее все элементы должны стать списками. НЕТ. все что я написал - пздежь!!! set это множество. оно должно выглядеть как список. 
                ## получить элемент множества можно только в цикле!!
                formula_shift_day = '=COUNT('
                
                for fday in formula_working_shift_day:
                    formula_shift_day = formula_shift_day + 'IF(SUM('
                    formula_working_shift_day[fday] = set(formula_working_shift_day[fday])
                    # print('formula_working_shift_day__string-->>', formula_working_shift_day[fday])
                    # print('formula_working_shift_day-->', formula_working_shift_day)
                    for i in formula_working_shift_day[fday]:
                        # print('i-->>', i)
                        formula_shift_day = formula_shift_day + i + ','
                    formula_shift_day = formula_shift_day.rstrip(',')
                    formula_shift_day = formula_shift_day + ')>0,1,"FALSE"),'
                    
                    # print('set-->>', set(formula_working_shift_day[fday]))
                # print('formula_working_shift_day-->>', formula_working_shift_day)
                formula_shift_day = formula_shift_day.rstrip(',')
                formula_shift_day = formula_shift_day + ')'
                # print('formula_shift_day-->>', formula_shift_day)
                if formula_shift_day == '=COUNT()':
                    formula_shift_day = ''
                ws.cell(row=lr_last, column=39).value = formula_shift_day
                ws.cell(row=lr_last, column=39).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                
                
                
                # -----------------ИТОГО отпуск--------------------------------
                ws.cell(row=lr_last, column=41).value = '=MAX('+ws.cell(row=start_lr_last+1, column=41).coordinate+':'+ws.cell(row=lr_last-1, column=41).coordinate+')'
                ws.cell(row=lr_last, column=41).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # -------------------------------------------------------------
                # -----------------ИТОГО отгул --------------------------------
                ws.cell(row=lr_last, column=42).value = '=MAX('+ws.cell(row=start_lr_last+1, column=42).coordinate+':'+ws.cell(row=lr_last-1, column=42).coordinate+')'
                ws.cell(row=lr_last, column=42).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # -------------------------------------------------------------
                # ------------------ИТОГО неявка-------------------------------
                ws.cell(row=lr_last, column=44).value = '=MAX('+ws.cell(row=start_lr_last+1, column=44).coordinate+':'+ws.cell(row=lr_last-1, column=44).coordinate+')'
                ws.cell(row=lr_last, column=44).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # -------------------------------------------------------------
                # ----------------ИТОГО больничный-----------------------------
                ws.cell(row=lr_last, column=43).value = '=MAX('+ws.cell(row=start_lr_last+1, column=43).coordinate+':'+ws.cell(row=lr_last-1, column=43).coordinate+')'
                ws.cell(row=lr_last, column=43).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # -------------------------------------------------------------
                
                # ------------------ИТОГО командировка-------------------------
                ws.cell(row=lr_last, column=45).value = '=SUM('+ws.cell(row=start_lr_last+1, column=45).coordinate+':'+ws.cell(row=lr_last-1, column=45).coordinate+')'
                ws.cell(row=lr_last, column=45).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # # -------------------------------------------------------------
                
                # # ----------------ИТОГО часы день------------------------------
                ws.cell(row=lr_last, column=46).value = '=SUM('+ws.cell(row=start_lr_last+1, column=46).coordinate+':'+ws.cell(row=lr_last-1, column=46).coordinate+')'
                ws.cell(row=lr_last, column=46).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # # -------------------------------------------------------------
                
                # # ---------------ИТОГО часы работа в выходые (день)------------
                ws.cell(row=lr_last, column=48).value = '=SUM('+ws.cell(row=start_lr_last+1, column=48).coordinate+':'+ws.cell(row=lr_last-1, column=48).coordinate+')'
                ws.cell(row=lr_last, column=48).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                
                
                
                
                formula_shift_night = '=COUNT('
                
                for fday in formula_working_shift_night:
                    formula_shift_night = formula_shift_night + 'IF(SUM('
                    formula_working_shift_night[fday] = set(formula_working_shift_night[fday])
                    # print('formula_working_shift_day__string-->>', formula_working_shift_night[fday])
                    for i in formula_working_shift_night[fday]:
                        # print('i-->>', i)
                        formula_shift_night = formula_shift_night + i + ','
                    formula_shift_night = formula_shift_night.rstrip(',')
                    formula_shift_night = formula_shift_night + ')>0,1,"FALSE"),'
                    
                    # print('set-->>', set(formula_working_shift_night[fday]))
                # print('formula_working_shift_night-->>', formula_working_shift_night)
                formula_shift_night = formula_shift_night.rstrip(',')
                formula_shift_night = formula_shift_night + ')'
                # print('formula_shift_night-->>', formula_shift_night)
                if formula_shift_night == '=COUNT()':
                    formula_shift_night = ''
                ws.cell(row=lr_last, column=40).value = formula_shift_night
                ws.cell(row=lr_last, column=40).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # ---------------------------------------------------------------------------------------------------------------------
                
                
                
                # # ---------------ИТОГО часы ночь-------------------------------
                ws.cell(row=lr_last, column=47).value = '=SUM('+ws.cell(row=start_lr_last+1, column=47).coordinate+':'+ws.cell(row=lr_last-1, column=47).coordinate+')'
                ws.cell(row=lr_last, column=47).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # # -------------------------------------------------------------
                # # ---------------ИТОГО часы работа в выходые (ночь)------------
                ws.cell(row=lr_last, column=49).value = '=SUM('+ws.cell(row=start_lr_last+1, column=49).coordinate+':'+ws.cell(row=lr_last-1, column=49).coordinate+')'
                ws.cell(row=lr_last, column=49).font = openpyxl.styles.Font(name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                # ---------------------------------------------------------------
                
                
                
                
# =============================================================================
#                 кусок от подсчета формул ИТОГО отработано смен день/ночь
                # flag_day = 'false'
                # flag_night = 'false'
                # formula_shift_day = '=COUNT('
                # formula_shift_night = '=COUNT('
                
# =============================================================================
                
                # print('days-->>', days)
                for day in days:
                    # print('days-->>', days)
                    # print('day-->>', day)
                    # print('df_total-->>', df_total)
                    # print('search--->>>', list(df_total['work_time'].loc[(df_total['job_name'] == list(job_description['job_name'].values())[0]) & (df_total['work_calendar_day'] == day['work_calendar_day'])])[0])
                    # print('search--->>>', len(df_total['work_time'].loc[(df_total['job_name'] == list(job_description['job_name'].values())[0]) & (df_total['work_calendar_day'] == day['work_calendar_day'])]))
                    
                    # !!!!!!!!!! тут хуйня. надо дополнить выборку подразделением и должностью, но это не точно
                    # !!! и вообще странно, что я здесь оперирую ФИО/должностью/званием, а не ID
                    # if len(df_total['work_time'].loc[(df_total['job_name'] == list(job_description['job_name'].values())[0]) & (df_total['work_calendar_day'] == day['work_calendar_day'])]) >0:
                    if len(df_total['work_time'].loc[(df_total['job_name'] == list(job_description['job_name'].values())[0]) & 
                                                     (df_total['job_department'] == list(job_description['job_department'].values())[0]) & 
                                                     (df_total['job_position'] == list(job_description['job_position'].values())[0]) & 
                                                     (df_total['work_calendar_day'] == day['work_calendar_day']) & 
                                                     (df_total['period_month'] == day['period_month'])]) >0:
                        
                        # print('-->>', df_total['work_time'].loc[(df_total['job_name'] == list(job_description['job_name'].values())[0]) & 
                        #                              (df_total['job_department'] == list(job_description['job_department'].values())[0]) & 
                        #                              (df_total['job_position'] == list(job_description['job_position'].values())[0]) & 
                        #                              (df_total['work_calendar_day'] == day['work_calendar_day']) & 
                        #                              (df_total['period_month'] == day['period_month'])])
                        # ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(df_total['work_time'].loc[(df_total['job_name'] == list(job_description['job_name'].values())[0]) & (df_total['work_calendar_day'] == day['work_calendar_day'])])[0])
                        ws.cell(row=lr_last, column=7+day['work_calendar_day'], value=list(df_total['work_time'].loc[(df_total['job_name'] == list(job_description['job_name'].values())[0]) & 
                                                                                                                     (df_total['job_department'] == list(job_description['job_department'].values())[0]) & 
                                                                                                                     (df_total['job_position'] == list(job_description['job_position'].values())[0]) & 
                                                                                                                     (df_total['work_calendar_day'] == day['work_calendar_day']) & 
                                                                                                                     (df_total['period_month'] == day['period_month'])
                                                                                                                     ])[0])
                    # кусок от подсчета формул ИТОГО отработано смен день/ночь
                    # formula_shift_day = formula_shift_day + 'IF(SUM('
                    # formula_shift_night = formula_shift_night + 'IF(SUM('
                    
# =============================================================================
                    # flag_day = 'false'
                    # flag_night = 'false'
                    # formula_shift_day = '=COUNT('
                    # formula_shift_night = '=COUNT('
                    # for s in range(6, len(array_total[0])):
                    #     formula_shift_day = formula_shift_day + 'IF(SUM('
                    #     formula_shift_night = formula_shift_night + 'IF(SUM('
                    #     for k in range(len(array_total)):
                    #         if array_total[k][5] == 'day':
                    #             flag_day = 'true'
                    #             formula_shift_day = formula_shift_day + \
                    #                 ws.cell(row=lr_last+k,
                    #                         column=s + 1).coordinate + ','
                    #         if array_total[k][5] == 'night':
                    #             flag_night = 'true'
                    #             formula_shift_night = formula_shift_night + \
                    #                 ws.cell(row=lr_last+k,
                    #                         column=s + 1).coordinate + ','
                    #     formula_shift_day = formula_shift_day.rstrip(',')
                    #     formula_shift_day = formula_shift_day + ')>0,1,"FALSE"),'
                    #     formula_shift_night = formula_shift_night.rstrip(',')
                    #     formula_shift_night = formula_shift_night + ')>0,1,"FALSE"),'
                    
                    # formula_shift_day = formula_shift_day.rstrip(',')
                    # formula_shift_day = formula_shift_day + ')'
                    # formula_shift_night = formula_shift_night.rstrip(',')
                    # formula_shift_night = formula_shift_night + ')'
                    # if flag_day == 'true':
                    #     ws.cell(row=last_empty_row+1,
                    #             column=39).value = formula_shift_day
                    #     ws.cell(row=last_empty_row+1, column=39).font = openpyxl.styles.Font(
                    #         name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    # if flag_night == 'true':
                    #     ws.cell(row=last_empty_row+1,
                    #             column=40).value = formula_shift_night
                    #     ws.cell(row=last_empty_row+1, column=40).font = openpyxl.styles.Font(
                    #         name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
# =============================================================================
                    
                    
                    
                for i in range(1,50):
                    ws.cell(row=lr_last, column=i).font = openpyxl.styles.Font(
                        name='Calibri', charset=204, family=2.0, b=True, sz=11.0)
                    ws.cell(row=lr_last, column=i).fill = PatternFill(
                        start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
#------------------------------------------------------------------------------  
def beauty(date_string, wb):
    ws = wb[date_string]
    thin_border1 = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    last_empty_row = len(list(ws.rows))+1
    for row in range(6, last_empty_row):
        for col in range(1, 50):
            if col == 1 and row > 6:
                ws.cell(row=row, column=1, value=row-6)
            ws.cell(row=row, column=col).border = thin_border1
#------------------------------------------------------------------------------
def create_exel(df_base, df_total, df_json, arg, config, flag):

    if flag == True: #если флаг равен 1, то мы будем дописывать в табель месяцы, которые пришли в json
        # print(config["excel"]["excel_path"] + config["excel"]["excel_name"] + ' ' + arg['worker'] + config["excel"][
        #     "typefile"])
        wb = load_workbook(config["excel"]["excel_path"]+config["excel"]["excel_name"]+' '+arg['worker']+config["excel"]["typefile"])

        logging.info(config["excel"]["excel_path"]+config["excel"]["excel_name"]+' '+arg['worker']+config["excel"]["typefile"])
        month_now = month_to_insert(df_json)
                
        #меняем порядок месяцев на нормальный
        month_now = list(month_now.values())
        months = ['','Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
        k = {}
        for i in range(len(months)):
            for j in month_now:
                if j == months[i]:
                    k[i] = j
        # print(k)
        month_now.clear()
        month_now = k
        # print(month_now)
        
    else: #если флаг равен 0, то мы создаем табель заного с начала года
        wb = Workbook()
        # удаялем все листы в экселе, созданные по умолчанию    
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            del wb[sheet.title]
        month_now = df_base['period_month'].drop_duplicates().to_dict()
        #month_now = {'0':'Август'}
    if arg['worker'] != 'all':
        id_worker = df_base[['job_name', 'job_department', 'job_position', 'id_worker']].loc[df_base['job_name'] == arg['worker']].drop_duplicates().to_dict()
             
    else:
    
        id_worker = df_base[['job_name', 'job_department', 'job_position', 'id_worker']].drop_duplicates().to_dict()
    logging.info("Start creating excel")
    for i in month_now.values():
        # print('fdsf')
        logging.info("Start insert " + i)
        insert_exel(config, df_total, wb, month_now = i, add = 'new', df_base=df_base, year_now=arg['year'])
        for j in list(id_worker['id_worker'].values()):
            insert_exel(config, df_total, wb, month_now = i, add = 'add', df_base=df_base, year_now=arg['year'], id_worker = j)
        beauty(i, wb)
        logging.info("Finish insert " + i)
    # wb.save(config["excel"]["excel_path"]+config["excel"]["excel_name"]+' '+arg['worker']+'_'+month_now + ' ' + year_now +config["excel"]["typefile"])
    logging.info("Finish creating excel")
    wb.save(config["excel"]["excel_path"]+config["excel"]["excel_name"]+' '+arg['worker']+config["excel"]["typefile"])
    
    